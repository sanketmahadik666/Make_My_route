"""
Simple data examination without pandas dependency
"""

import os
import sys
from pathlib import Path
import openpyxl

def examine_excel_structure_simple(file_path):
    """Examine the structure of an Excel file using openpyxl."""
    print(f"\n{'='*60}")
    print(f"EXAMINING: {file_path}")
    print(f"{'='*60}")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        print(f"Sheet names: {wb.sheetnames}")
        
        # Examine each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            
            try:
                ws = wb[sheet_name]
                
                # Get dimensions
                max_row = ws.max_row
                max_col = ws.max_column
                print(f"Dimensions: {max_row} rows × {max_col} columns")
                
                # Get column headers (first row)
                headers = []
                for col in range(1, min(max_col + 1, 50)):  # Limit to first 50 columns
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
                    else:
                        headers.append(f"Column_{col}")
                
                print(f"Columns ({len(headers)}):")
                for i, header in enumerate(headers, 1):
                    print(f"  {i:2d}. {header}")
                
                # Show first few data rows
                print(f"\nFirst 3 data rows:")
                for row in range(2, min(5, max_row + 1)):
                    row_data = []
                    for col in range(1, min(len(headers) + 1, 10)):  # Show first 10 columns
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is not None:
                            row_data.append(str(cell_value)[:20])  # Truncate long values
                        else:
                            row_data.append("NULL")
                    print(f"  Row {row}: {' | '.join(row_data)}")
                
            except Exception as e:
                print(f"Error reading sheet {sheet_name}: {e}")
    
    except Exception as e:
        print(f"Error examining file: {e}")

def main():
    """Main function to examine all CS2_35 files."""
    data_dir = Path("/home/sanket/Make_My_route/DATA/CS2_35")
    
    if not data_dir.exists():
        print(f"Data directory not found: {data_dir}")
        return
    
    # Get all Excel files
    excel_files = list(data_dir.glob("*.xlsx"))
    excel_files.sort()
    
    print(f"Found {len(excel_files)} Excel files")
    
    # Examine first 3 files in detail
    for i, file_path in enumerate(excel_files[:3]):
        examine_excel_structure_simple(file_path)
        
        if i < len(excel_files) - 1:
            print(f"\n{'#'*60}")
    
    # Quick overview of all files
    print(f"\n{'='*60}")
    print("QUICK OVERVIEW - ALL FILES")
    print(f"{'='*60}")
    
    for file_path in excel_files:
        try:
            wb = openpyxl.load_workbook(file_path)
            print(f"{file_path.name}: {wb.sheetnames}")
        except Exception as e:
            print(f"{file_path.name}: ERROR - {e}")

if __name__ == "__main__":
    main()